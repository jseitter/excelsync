"""
Tests for ExcelSync class.
"""

import os
import json
import tempfile
from pathlib import Path

import pytest
import openpyxl

from excelsync import ExcelSync


@pytest.fixture
def sample_excel_file():
    """Create a sample Excel file for testing."""
    # Create a temporary Excel file
    with tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False) as temp:
        temp_path = temp.name
    
    # Create a workbook with test data
    wb = openpyxl.Workbook()
    
    # Add data to default sheet
    sheet = wb.active
    sheet.title = "Sheet1"
    
    # Add headers
    sheet["A1"] = "ID"
    sheet["B1"] = "Name"
    sheet["C1"] = "Age"
    sheet["D1"] = "Date"
    
    # Add data
    sheet["A2"] = 1
    sheet["B2"] = "John Doe"
    sheet["C2"] = 30
    sheet["D2"] = "2023-01-01"
    
    sheet["A3"] = 2
    sheet["B3"] = "Jane Smith"
    sheet["C3"] = 28
    sheet["D3"] = "2023-02-15"
    
    # Add a second sheet
    sheet2 = wb.create_sheet(title="Sheet2")
    sheet2["A1"] = "Product"
    sheet2["B1"] = "Price"
    sheet2["A2"] = "Widget"
    sheet2["B2"] = 19.99
    
    # Save the workbook
    wb.save(temp_path)
    
    yield temp_path
    
    # Clean up
    if os.path.exists(temp_path):
        os.unlink(temp_path)


@pytest.fixture
def custom_header_row_excel_file():
    """Create a sample Excel file with headers in row 3."""
    # Create a temporary Excel file
    with tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False) as temp:
        temp_path = temp.name
    
    # Create a workbook with test data
    wb = openpyxl.Workbook()
    
    # Add data to default sheet
    sheet = wb.active
    sheet.title = "CustomHeaderSheet"
    
    # Add some title or metadata in rows 1-2
    sheet["A1"] = "EXAMPLE COMPANY"
    sheet["A2"] = "Employee Database"
    
    # Add headers in row 3
    sheet["A3"] = "ID"
    sheet["B3"] = "Name"
    sheet["C3"] = "Age"
    sheet["D3"] = "Date"
    
    # Add data starting from row 4
    sheet["A4"] = 1
    sheet["B4"] = "John Doe"
    sheet["C4"] = 30
    sheet["D4"] = "2023-01-01"
    
    sheet["A5"] = 2
    sheet["B5"] = "Jane Smith"
    sheet["C5"] = 28
    sheet["D5"] = "2023-02-15"
    
    # Save the workbook
    wb.save(temp_path)
    
    yield temp_path
    
    # Clean up
    if os.path.exists(temp_path):
        os.unlink(temp_path)


def test_init(sample_excel_file):
    """Test initialization of ExcelSync."""
    excel_sync = ExcelSync(sample_excel_file)
    assert excel_sync.excel_file.exists()
    assert len(excel_sync.workbook.sheetnames) == 2
    assert "Sheet1" in excel_sync.workbook.sheetnames
    assert "Sheet2" in excel_sync.workbook.sheetnames
    assert excel_sync.header_row == 1  # Default header row


def test_extract_structure(sample_excel_file):
    """Test extracting structure from Excel file."""
    excel_sync = ExcelSync(sample_excel_file)
    structure = excel_sync.extract_structure()
    
    assert "sheets" in structure
    assert "Sheet1" in structure["sheets"]
    assert "Sheet2" in structure["sheets"]
    
    sheet1 = structure["sheets"]["Sheet1"]
    assert "headers" in sheet1
    assert len(sheet1["headers"]) == 4
    assert sheet1["headers"][1]["name"] == "ID"
    assert sheet1["headers"][2]["name"] == "Name"
    assert sheet1["headers"][3]["name"] == "Age"
    assert sheet1["headers"][4]["name"] == "Date"
    
    # Check data types
    assert sheet1["headers"][1]["data_type"] == "integer"
    assert sheet1["headers"][2]["data_type"] == "string"
    assert sheet1["headers"][3]["data_type"] == "integer"


def test_export_structure(sample_excel_file):
    """Test exporting structure to a file."""
    excel_sync = ExcelSync(sample_excel_file)
    
    with tempfile.NamedTemporaryFile(suffix='.json', delete=False) as temp:
        temp_path = temp.name
    
    excel_sync.export_structure(temp_path)
    
    # Verify the exported structure
    assert os.path.exists(temp_path)
    with open(temp_path, 'r') as f:
        exported_structure = json.load(f)
    
    assert "sheets" in exported_structure
    assert "Sheet1" in exported_structure["sheets"]
    assert "Sheet2" in exported_structure["sheets"]
    
    # Clean up
    os.unlink(temp_path)


def test_validate_structure(sample_excel_file):
    """Test validating Excel structure."""
    excel_sync = ExcelSync(sample_excel_file)
    structure = excel_sync.extract_structure()
    
    # Structure should be valid against itself
    is_valid, issues = excel_sync.validate_structure(structure)
    assert is_valid
    assert len(issues) == 0
    
    # Modify structure to create an invalid one
    invalid_structure = structure.copy()
    invalid_structure["sheets"]["Sheet1"]["headers"][1]["name"] = "Modified"
    
    is_valid, issues = excel_sync.validate_structure(invalid_structure)
    assert not is_valid
    assert len(issues) > 0


def test_export_to_yaml(sample_excel_file):
    """Test exporting Excel data to YAML."""
    excel_sync = ExcelSync(sample_excel_file)
    
    with tempfile.NamedTemporaryFile(suffix='.yaml', delete=False) as temp:
        temp_path = temp.name
    
    excel_sync.export_to_yaml(temp_path)
    
    # Verify the file exists
    assert os.path.exists(temp_path)
    
    # Clean up
    os.unlink(temp_path)


def test_custom_header_row(custom_header_row_excel_file):
    """Test using a custom header row."""
    # Initialize with header_row=3
    excel_sync = ExcelSync(custom_header_row_excel_file, header_row=3)
    structure = excel_sync.extract_structure()
    
    assert "sheets" in structure
    assert "CustomHeaderSheet" in structure["sheets"]
    assert structure["file_properties"]["header_row"] == 3
    
    sheet = structure["sheets"]["CustomHeaderSheet"]
    assert "headers" in sheet
    assert len(sheet["headers"]) == 4
    assert sheet["headers"][1]["name"] == "ID"
    assert sheet["headers"][2]["name"] == "Name"
    assert sheet["headers"][3]["name"] == "Age"
    assert sheet["headers"][4]["name"] == "Date"
    
    # Verify data types are detected from row 4 (header_row + 1)
    assert sheet["headers"][1]["data_type"] == "integer"
    assert sheet["headers"][2]["data_type"] == "string"
    assert sheet["headers"][3]["data_type"] == "integer"


def test_override_header_row(custom_header_row_excel_file):
    """Test overriding the header row in method calls."""
    # Initialize with default header_row=1
    excel_sync = ExcelSync(custom_header_row_excel_file)
    
    # But extract structure with header_row=3
    structure = excel_sync.extract_structure(header_row=3)
    
    assert structure["file_properties"]["header_row"] == 3
    sheet = structure["sheets"]["CustomHeaderSheet"]
    assert "headers" in sheet
    assert len(sheet["headers"]) == 4
    assert sheet["headers"][1]["name"] == "ID"
    assert sheet["headers"][2]["name"] == "Name"
    
    # Check that the instance's header_row is still 1
    assert excel_sync.header_row == 1
    
    # Export to YAML with header_row=3
    with tempfile.NamedTemporaryFile(suffix='.yaml', delete=False) as temp:
        temp_path = temp.name
    
    excel_sync.export_to_yaml(temp_path, header_row=3)
    
    # Verify the file exists
    assert os.path.exists(temp_path)
    
    # Clean up
    os.unlink(temp_path)


def test_header_row_validation(custom_header_row_excel_file):
    """Test validation with different header rows."""
    excel_sync = ExcelSync(custom_header_row_excel_file, header_row=3)
    structure = excel_sync.extract_structure()
    
    # Structure should be valid against itself
    is_valid, issues = excel_sync.validate_structure(structure)
    assert is_valid
    assert len(issues) == 0
    
    # Modify the header row and validate
    modified_structure = structure.copy()
    modified_structure["file_properties"]["header_row"] = 4
    
    is_valid, issues = excel_sync.validate_structure(modified_structure)
    assert not is_valid
    assert len(issues) > 0
    assert any("Header row mismatch" in issue for issue in issues)


def test_save_and_load_structure_with_header_row(custom_header_row_excel_file):
    """Test saving and loading structure with header row information."""
    excel_sync = ExcelSync(custom_header_row_excel_file, header_row=3)
    
    with tempfile.NamedTemporaryFile(suffix='.json', delete=False) as temp:
        temp_path = temp.name
    
    excel_sync.export_structure(temp_path)
    
    # Create a new instance with default header_row=1
    new_excel_sync = ExcelSync(custom_header_row_excel_file)
    assert new_excel_sync.header_row == 1
    
    # Load the structure, which should update the header_row
    new_excel_sync.load_structure(temp_path)
    assert new_excel_sync.header_row == 3
    
    # Clean up
    os.unlink(temp_path) 