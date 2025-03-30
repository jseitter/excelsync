"""
Simple example demonstrating the usage of ExcelSync.
"""

import os
import sys
import tempfile
from pathlib import Path

import openpyxl

# Add the parent directory to the path so we can import excelsync
sys.path.insert(0, str(Path(__file__).parent.parent))
from excelsync import ExcelSync


def create_sample_excel():
    """Create a sample Excel file for demonstration."""
    with tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False) as temp:
        temp_path = temp.name
    
    print(f"Creating sample Excel file at: {temp_path}")
    
    # Create a workbook with test data
    wb = openpyxl.Workbook()
    
    # Add data to default sheet
    sheet = wb.active
    sheet.title = "Employees"
    
    # Add headers
    sheet["A1"] = "Employee ID"
    sheet["B1"] = "Name"
    sheet["C1"] = "Department"
    sheet["D1"] = "Salary"
    sheet["E1"] = "Start Date"
    
    # Add data
    data = [
        (101, "John Smith", "Engineering", 75000, "2022-01-15"),
        (102, "Lisa Johnson", "Marketing", 68000, "2021-05-20"),
        (103, "Robert Garcia", "Finance", 82000, "2023-03-10"),
        (104, "Sarah Chen", "Engineering", 78000, "2022-09-05"),
        (105, "Michael Williams", "HR", 65000, "2021-11-28"),
    ]
    
    for row_idx, row_data in enumerate(data, start=2):
        for col_idx, value in enumerate(row_data, start=1):
            sheet.cell(row=row_idx, column=col_idx, value=value)
    
    # Add a second sheet
    sheet2 = wb.create_sheet(title="Departments")
    sheet2["A1"] = "Department"
    sheet2["B1"] = "Manager"
    sheet2["C1"] = "Budget"
    
    dept_data = [
        ("Engineering", "Alex Johnson", 1200000),
        ("Marketing", "Emily Davis", 800000),
        ("Finance", "David Wilson", 950000),
        ("HR", "Jessica Martinez", 500000),
    ]
    
    for row_idx, row_data in enumerate(dept_data, start=2):
        for col_idx, value in enumerate(row_data, start=1):
            sheet2.cell(row=row_idx, column=col_idx, value=value)
    
    # Save the workbook
    wb.save(temp_path)
    return temp_path


def main():
    """Run the demonstration."""
    # Create sample Excel file
    excel_path = create_sample_excel()
    
    # Initialize ExcelSync with the sample file
    print("\nInitializing ExcelSync...")
    excel_sync = ExcelSync(excel_path)
    
    # Extract and print structure
    print("\nExtracting structure...")
    structure = excel_sync.extract_structure()
    print(f"Found {len(structure['sheets'])} sheets: {', '.join(structure['sheets'].keys())}")
    
    # Export structure to file
    structure_file = "sample_structure.json"
    print(f"\nExporting structure to {structure_file}...")
    excel_sync.export_structure(structure_file)
    print(f"Structure exported to {structure_file}")
    
    # Validate structure
    print("\nValidating structure...")
    is_valid, issues = excel_sync.validate_structure(structure)
    print(f"Structure is valid: {is_valid}")
    if not is_valid:
        print("Issues found:")
        for issue in issues:
            print(f"- {issue}")
    
    # Export to YAML
    yaml_file = "sample_data.yaml"
    print(f"\nExporting data to YAML ({yaml_file})...")
    excel_sync.export_to_yaml(yaml_file)
    print(f"Data exported to {yaml_file}")
    
    # Cleanup
    print("\nCleaning up...")
    os.unlink(excel_path)
    print("Sample Excel file deleted")
    print("Done!")


if __name__ == "__main__":
    main() 